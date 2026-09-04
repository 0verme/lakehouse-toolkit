import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from shared.lineage.asset_tables import AssetMappingLoadError
from tools.misc.xlsx_sql_tables import (
    ASSET_CHECK_FAILED,
    ASSET_HEADER,
    LEGACY_TABLES_HEADER,
    SQL_HEADER,
    STATUS_EMPTY,
    STATUS_HEADER,
    STATUS_INVALID,
    STATUS_SUCCESS,
    STATUS_UNRECOGNIZED,
    TABLES_HEADER,
    build_output_filename,
    extract_sql_tables,
    parse_workbook,
    parse_workbook_with_asset_lookup,
)


def workbook_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def build_test_workbook():
    workbook = Workbook()
    target = workbook.active
    target.title = "SQL清单"
    target.append(["编号", "名称", "备注", f"  {SQL_HEADER}  "])
    target.append(
        [
            1,
            "示例一",
            "保留",
            "insert into DWM.M_DEMO_RESULT select * from DEMO_META.JOBS j join DWF.F_DEMO_EVENT d on j.id=d.id",
        ]
    )
    target.append(
        [2, "示例二", "保留", "select *\nfrom dwa.a_customer using DWD.D_ACCOUNT"]
    )
    target.append([3, "空 SQL", "保留", None])
    target.append([4, "未识别", "保留", "select current_date"])
    target.append([5, "格式异常", "保留", 12345])
    target.append([None, None, None, None])
    target["C2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.create_sheet("不处理").append(
        [SQL_HEADER, "select * from DEMO_META.OTHER_TABLE"]
    )
    return workbook


class XlsxSqlTablesTests(unittest.TestCase):
    def test_extract_sql_tables_ignores_comments_literals_and_ctes(self):
        sql = """
        -- select * from DEMO_META.COMMENT_TABLE
        with recent as (
            select * from "demo_meta"."jobs"
        ), second_cte as (
            select * from recent join [DWF].[F_DEMO_EVENT] d on 1=1
        )
        select 'from DWA.STRING_TABLE' from second_cte
        """

        self.assertEqual(
            extract_sql_tables(sql), ["DEMO_META.JOBS", "DWF.F_DEMO_EVENT"]
        )

    def test_extract_sql_tables_supports_dml_ddl_and_quoted_identifiers(self):
        sql = """
        insert overwrite table `dwm`.`m_result` select * from dwa.a_source;
        update "DWD"."D_ACCOUNT" set flag = 1;
        merge into [DWF].[F_TARGET] t using DWO.O_SOURCE s on 1=1;
        delete from DWP.P_DELETE;
        create table if not exists DM.M_CREATE (id int);
        alter table DM.M_ALTER add column name varchar(10);
        truncate table DM.M_TRUNCATE;
        drop table if exists DM.M_DROP;
        """

        self.assertEqual(
            extract_sql_tables(sql),
            [
                "DM.M_ALTER",
                "DM.M_CREATE",
                "DM.M_DROP",
                "DM.M_TRUNCATE",
                "DWA.A_SOURCE",
                "DWD.D_ACCOUNT",
                "DWF.F_TARGET",
                "DWM.M_RESULT",
                "DWO.O_SOURCE",
                "DWP.P_DELETE",
            ],
        )

    def test_parse_workbook_uses_first_sheet_header_and_writes_statuses(self):
        result = parse_workbook(
            workbook_bytes(build_test_workbook()),
            asset_plan_map={
                "DWF.F_DEMO_EVENT": "DEMO_PLAN_INGEST_DAY",
                "DEMO_META.JOBS": "DEMO_PLAN_EXPORT_DAY",
            },
        )
        parsed = load_workbook(BytesIO(result.content))
        target = parsed.worksheets[0]

        headers = [cell.value for cell in target[1]]
        tables_column = headers.index(TABLES_HEADER) + 1
        status_column = headers.index(STATUS_HEADER) + 1
        asset_column = headers.index(ASSET_HEADER) + 1
        self.assertNotEqual(tables_column, status_column)
        self.assertEqual(result.sheet_title, "SQL清单")
        self.assertEqual(result.successful_rows, 2)
        self.assertEqual(result.unrecognized_rows, 1)
        self.assertEqual(result.empty_rows, 1)
        self.assertEqual(result.invalid_rows, 1)
        self.assertEqual(result.asset_related_rows, 1)
        self.assertFalse(result.asset_check_failed)
        self.assertEqual(
            target.cell(row=2, column=tables_column).value,
            "DEMO_META.JOBS\nDWF.F_DEMO_EVENT\nDWM.M_DEMO_RESULT",
        )
        self.assertEqual(target.cell(row=2, column=status_column).value, STATUS_SUCCESS)
        self.assertEqual(target.cell(row=4, column=status_column).value, STATUS_EMPTY)
        self.assertEqual(
            target.cell(row=5, column=status_column).value, STATUS_UNRECOGNIZED
        )
        self.assertEqual(target.cell(row=6, column=status_column).value, STATUS_INVALID)
        self.assertEqual(
            target.cell(row=2, column=asset_column).value, "demo_export\ndemo_ingest"
        )
        self.assertEqual(
            target.cell(row=2, column=asset_column).fill.fgColor.rgb, "00FCE8E6"
        )
        self.assertEqual(
            target.cell(row=2, column=asset_column).font.color.rgb, "00C00000"
        )
        self.assertIsNone(target.cell(row=3, column=asset_column).value)
        self.assertIsNone(target.cell(row=7, column=status_column).value)
        self.assertEqual(target["C2"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(parsed.worksheets[1].max_column, 2)
        parsed.close()

    def test_parse_workbook_reuses_output_columns_and_migrates_legacy_header(self):
        workbook = build_test_workbook()
        sheet = workbook.worksheets[0]
        sheet.cell(row=1, column=5, value=LEGACY_TABLES_HEADER)
        first_result = parse_workbook(workbook_bytes(workbook))
        second_result = parse_workbook(first_result.content)
        parsed = load_workbook(BytesIO(second_result.content))
        headers = [cell.value for cell in parsed.worksheets[0][1]]

        self.assertEqual(headers.count(TABLES_HEADER), 1)
        self.assertEqual(headers.count(STATUS_HEADER), 1)
        self.assertEqual(headers.count(ASSET_HEADER), 1)
        self.assertNotIn(LEGACY_TABLES_HEADER, headers)
        parsed.close()

    def test_parse_workbook_marks_asset_check_failure_without_blocking_export(self):
        result = parse_workbook(
            workbook_bytes(build_test_workbook()),
            asset_check_failed=True,
        )
        parsed = load_workbook(BytesIO(result.content))
        sheet = parsed.active
        headers = [cell.value for cell in sheet[1]]
        asset_column = headers.index(ASSET_HEADER) + 1

        self.assertTrue(result.asset_check_failed)
        self.assertEqual(result.asset_related_rows, 0)
        self.assertEqual(
            sheet.cell(row=2, column=asset_column).value, ASSET_CHECK_FAILED
        )
        self.assertEqual(
            sheet.cell(row=2, column=asset_column).fill.fgColor.rgb, "00FFF4CC"
        )
        self.assertEqual(
            sheet.cell(row=2, column=asset_column).font.color.rgb, "009C6500"
        )
        self.assertIsNone(sheet.cell(row=7, column=asset_column).value)
        parsed.close()

    def test_parse_workbook_clears_stale_asset_highlight_when_mapping_changes(self):
        first_result = parse_workbook(
            workbook_bytes(build_test_workbook()),
            asset_plan_map={
                "DWF.F_DEMO_EVENT": "DEMO_PLAN_INGEST_DAY",
            },
        )
        second_result = parse_workbook(first_result.content, asset_plan_map={})
        parsed = load_workbook(BytesIO(second_result.content))
        sheet = parsed.active
        headers = [cell.value for cell in sheet[1]]
        asset_cell = sheet.cell(row=2, column=headers.index(ASSET_HEADER) + 1)

        self.assertIsNone(asset_cell.value)
        self.assertIsNone(asset_cell.fill.fill_type)
        self.assertIsNone(asset_cell.font.color)
        parsed.close()

    @patch("tools.misc.xlsx_sql_tables.load_asset_plan_map")
    def test_asset_lookup_loads_mapping_once_per_workbook(self, load_mapping):
        load_mapping.return_value = {
            "DWF.F_DEMO_EVENT": "DEMO_PLAN_INGEST_DAY",
        }

        result = parse_workbook_with_asset_lookup(workbook_bytes(build_test_workbook()))

        self.assertEqual(result.asset_related_rows, 1)
        load_mapping.assert_called_once_with()

    @patch(
        "tools.misc.xlsx_sql_tables.load_asset_plan_map",
        side_effect=AssetMappingLoadError("db unavailable"),
    )
    def test_asset_lookup_falls_back_to_failure_markers(self, load_mapping):
        result = parse_workbook_with_asset_lookup(workbook_bytes(build_test_workbook()))

        self.assertTrue(result.asset_check_failed)
        load_mapping.assert_called_once_with()

    def test_parse_workbook_rejects_missing_sql_header(self):
        workbook = Workbook()
        workbook.active.append(["编号", "SQL"])

        with self.assertRaisesRegex(ValueError, "未找到“SQL脚本”表头"):
            parse_workbook(workbook_bytes(workbook))

    def test_parse_workbook_rejects_duplicate_sql_headers(self):
        workbook = Workbook()
        workbook.active.append([SQL_HEADER, f" {SQL_HEADER} "])

        with self.assertRaisesRegex(ValueError, "重复“SQL脚本”表头"):
            parse_workbook(workbook_bytes(workbook))

    def test_parse_workbook_marks_formula_as_invalid_and_preserves_formula(self):
        workbook = Workbook()
        workbook.active.append(["编号", SQL_HEADER])
        workbook.active.append([1, "=A2"])
        result = parse_workbook(workbook_bytes(workbook))
        parsed = load_workbook(BytesIO(result.content), data_only=False)
        sheet = parsed.active
        headers = [cell.value for cell in sheet[1]]

        self.assertEqual(sheet["B2"].value, "=A2")
        self.assertEqual(
            sheet.cell(row=2, column=headers.index(STATUS_HEADER) + 1).value,
            STATUS_INVALID,
        )
        parsed.close()

    def test_build_output_filename(self):
        self.assertEqual(build_output_filename("a.xlsx"), "a_SQL表名解析.xlsx")


if __name__ == "__main__":
    unittest.main()
