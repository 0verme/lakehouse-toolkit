from __future__ import annotations

import sqlite3
from collections import deque
from datetime import datetime
from importlib import import_module
from pathlib import Path

from openpyxl import load_workbook

from shared.config.metadata import table as metadata_table

ROOT_DIR = Path(__file__).resolve().parents[2]
MAPPING_XLSX_PATH = ROOT_DIR / "resources" / "xlsx" / "mapping.xlsx"
MAPPING_DB_PATH = ROOT_DIR / "runtime" / "sqlite" / "mapping_lineage.db"

HEADER_ALIASES = {
    "target_system": {"目标系统"},
    "target_schema": {"目标模式"},
    "target_table": {"目标表"},
    "target_column": {"目标字段"},
    "source_system": {"源系统"},
    "source_schema": {"源模式"},
    "source_table": {"源表"},
    "source_column": {"源字段"},
}


def normalize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "")


def normalize_token(value) -> str:
    return normalize_value(value).upper().replace(" ", "")


def parse_input_table_name(table_name: str) -> tuple[str, str]:
    normalized = normalize_token(table_name)
    if "." in normalized:
        schema, table = normalized.split(".", 1)
        return schema, table
    return "", normalized


def normalize_identifier(schema: str, table: str, column: str) -> tuple[str, str, str]:
    normalized_schema = normalize_token(schema)
    normalized_table = normalize_token(table)
    if not normalized_schema and "." in normalized_table:
        normalized_schema, normalized_table = normalized_table.split(".", 1)
    return normalized_schema, normalized_table, normalize_token(column)


def compact_identifier(identifier: tuple[str, str, str]) -> str:
    schema, table, column = identifier
    table_name = f"{schema}.{table}" if schema else table
    return f"{table_name}.{column}"


def normalize_registered_table_name(table_name: str) -> str:
    clean_name = normalize_value(table_name).split(" ")[0]
    return normalize_token(clean_name)


def detect_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True), start=1
    ):
        mapping = {}
        for col_index, value in enumerate(row, start=1):
            cell_text = normalize_value(value)
            for field_name, aliases in HEADER_ALIASES.items():
                if cell_text in aliases:
                    mapping[field_name] = col_index
        if (
            "target_table" in mapping
            and "target_column" in mapping
            and "source_table" in mapping
            and "source_column" in mapping
        ):
            return row_index, mapping
    raise ValueError("未在 Excel 中识别到“目标表/目标字段/源表/源字段”表头。")


def load_lineage_edges_from_xlsx(
    xlsx_path: str | Path = MAPPING_XLSX_PATH,
) -> list[dict]:
    workbook = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        last_error = None
        for worksheet in workbook.worksheets:
            try:
                header_row, columns = detect_header_row(worksheet)
            except ValueError as exc:
                last_error = exc
                continue
            rows = []
            for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                source_table = normalize_value(
                    values[columns["source_table"] - 1]
                    if columns.get("source_table")
                    else ""
                )
                source_column = normalize_value(
                    values[columns["source_column"] - 1]
                    if columns.get("source_column")
                    else ""
                )
                target_table = normalize_value(
                    values[columns["target_table"] - 1]
                    if columns.get("target_table")
                    else ""
                )
                target_column = normalize_value(
                    values[columns["target_column"] - 1]
                    if columns.get("target_column")
                    else ""
                )
                if not (
                    source_table and source_column and target_table and target_column
                ):
                    continue
                rows.append(
                    {
                        "source_system": normalize_value(
                            values[columns["source_system"] - 1]
                            if columns.get("source_system")
                            else ""
                        ),
                        "source_schema": normalize_value(
                            values[columns["source_schema"] - 1]
                            if columns.get("source_schema")
                            else ""
                        ),
                        "source_table": source_table,
                        "source_column": source_column,
                        "target_system": normalize_value(
                            values[columns["target_system"] - 1]
                            if columns.get("target_system")
                            else ""
                        ),
                        "target_schema": normalize_value(
                            values[columns["target_schema"] - 1]
                            if columns.get("target_schema")
                            else ""
                        ),
                        "target_table": target_table,
                        "target_column": target_column,
                    }
                )
            return rows
        raise last_error or ValueError("未在 Excel 中识别到有效血缘数据。")
    finally:
        workbook.close()


def ensure_db_parent(db_path: str | Path = MAPPING_DB_PATH):
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)


def recreate_mapping_sqlite(
    xlsx_path: str | Path = MAPPING_XLSX_PATH, db_path: str | Path = MAPPING_DB_PATH
) -> dict:
    xlsx_path = Path(xlsx_path)
    db_path = Path(db_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {xlsx_path}")

    ensure_db_parent(db_path)
    rows = load_lineage_edges_from_xlsx(xlsx_path)

    with sqlite3.connect(db_path) as conn:
        conn.execute("DROP TABLE IF EXISTS lineage_edge")
        conn.execute("DROP TABLE IF EXISTS lineage_meta")
        conn.execute(
            """
            CREATE TABLE lineage_edge (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_system TEXT NOT NULL,
                source_schema TEXT NOT NULL,
                source_table TEXT NOT NULL,
                source_column TEXT NOT NULL,
                target_system TEXT NOT NULL,
                target_schema TEXT NOT NULL,
                target_table TEXT NOT NULL,
                target_column TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE lineage_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )

        insert_rows = []
        for row in rows:
            source_schema, source_table, source_column = normalize_identifier(
                row["source_schema"], row["source_table"], row["source_column"]
            )
            target_schema, target_table, target_column = normalize_identifier(
                row["target_schema"], row["target_table"], row["target_column"]
            )
            insert_rows.append(
                (
                    normalize_token(row["source_system"]),
                    source_schema,
                    source_table,
                    source_column,
                    normalize_token(row["target_system"]),
                    target_schema,
                    target_table,
                    target_column,
                )
            )

        # pi-lens-ignore: python-sql-injection
        conn.executemany(
            """
            INSERT INTO lineage_edge (
                source_system, source_schema, source_table, source_column,
                target_system, target_schema, target_table, target_column
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            insert_rows,
        )
        conn.execute(
            "CREATE INDEX idx_lineage_source ON lineage_edge(source_schema, source_table, source_column)"
        )
        conn.execute(
            "CREATE INDEX idx_lineage_target ON lineage_edge(target_schema, target_table, target_column)"
        )

        xlsx_stat = xlsx_path.stat()
        meta_items = {
            "source_xlsx_path": str(xlsx_path),
            "source_xlsx_mtime_ns": str(xlsx_stat.st_mtime_ns),
            "source_xlsx_size": str(xlsx_stat.st_size),
            "imported_at": datetime.now().isoformat(timespec="seconds"),
            "edge_count": str(len(insert_rows)),
        }
        # pi-lens-ignore: python-sql-injection
        conn.executemany(
            "INSERT INTO lineage_meta(key, value) VALUES(?, ?)",
            list(meta_items.items()),
        )
        conn.commit()

    return {
        "db_path": str(db_path),
        "xlsx_path": str(xlsx_path),
        "edge_count": len(rows),
    }


def load_mapping_meta(db_path: str | Path = MAPPING_DB_PATH) -> dict[str, str]:
    db_path = Path(db_path)
    if not db_path.exists():
        return {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute("SELECT key, value FROM lineage_meta").fetchall()
    return dict(rows)


def get_mapping_db_status(
    db_path: str | Path = MAPPING_DB_PATH, xlsx_path: str | Path = MAPPING_XLSX_PATH
) -> dict:
    db_path = Path(db_path)
    xlsx_path = Path(xlsx_path)
    meta = load_mapping_meta(db_path)
    db_exists = db_path.exists()
    xlsx_exists = xlsx_path.exists()
    is_fresh = False
    if db_exists and xlsx_exists and meta:
        xlsx_stat = xlsx_path.stat()
        is_fresh = meta.get("source_xlsx_mtime_ns") == str(
            xlsx_stat.st_mtime_ns
        ) and meta.get("source_xlsx_size") == str(xlsx_stat.st_size)
    return {
        "db_exists": db_exists,
        "xlsx_exists": xlsx_exists,
        "is_fresh": is_fresh,
        "db_path": str(db_path),
        "xlsx_path": str(xlsx_path),
        "meta": meta,
    }


def _select_runtime_rows(profile: str, sql: str):
    if profile in ("demo_local", "local_pg"):
        try:
            profile_module = import_module("services.db_profile")
        except ModuleNotFoundError:
            profile_module = import_module("apps.svn_check.services.db_profile")

        if profile_module.is_postgres_profile():
            from shared.db.postgres import fetch_all

            return fetch_all(sql, profile=profile_module.get_active_audit_profile())
    from shared.db.gaussdb import select_sql_with_profile

    return select_sql_with_profile(profile, sql)


def load_registered_result_tables(profile: str = "demo_local") -> set[str]:
    sql = """
        SELECT target_table
        FROM __PROGRAMS_TABLE__
        WHERE target_table IS NOT NULL
        """.replace("__PROGRAMS_TABLE__", metadata_table("programs", "programs"))
    rows = _select_runtime_rows(profile, sql) or []
    result_tables = set()
    for row in rows:
        table_name = normalize_value(row[0]) if row else ""
        normalized_table_name = normalize_registered_table_name(table_name)
        if normalized_table_name:
            result_tables.add(normalized_table_name)
    return result_tables


def filter_registered_result_nodes(
    nodes: list[tuple[str, str, str]],
    result_tables: set[str] | None = None,
    profile: str = "demo_local",
) -> list[tuple[str, str, str]]:
    result_tables = (
        result_tables
        if result_tables is not None
        else load_registered_result_tables(profile)
    )
    filtered_nodes = []
    for schema, table, column in nodes:
        full_name = normalize_registered_table_name(
            f"{schema}.{table}" if schema else table
        )
        if full_name in result_tables:
            filtered_nodes.append((schema, table, column))
    return filtered_nodes


def find_start_nodes_in_sqlite(
    table_name: str, column_name: str, db_path: str | Path = MAPPING_DB_PATH
) -> list[tuple[str, str, str]]:
    schema, table = parse_input_table_name(table_name)
    column = normalize_token(column_name)
    query = """
        SELECT DISTINCT source_schema, source_table, source_column
        FROM lineage_edge
        WHERE source_table = ?
          AND source_column = ?
          AND (? IS NULL OR source_schema = ?)
    """
    params = (table, column, schema, schema)
    with sqlite3.connect(db_path) as conn:
        # pi-lens-ignore: python-sql-injection
        rows = conn.execute(query, params).fetchall()
    return [(row[0], row[1], row[2]) for row in rows]


def walk_downstream_in_sqlite(
    start_nodes: list[tuple[str, str, str]],
    db_path: str | Path = MAPPING_DB_PATH,
    max_depth: int | None = None,
) -> tuple[list[list[str]], list[tuple[str, str, str]]]:
    queue = deque()
    visited = set()
    downstream_nodes = set()
    relation_rows = []

    with sqlite3.connect(db_path) as conn:
        for start in start_nodes:
            queue.append((start, 0))
            visited.add(start)

        while queue:
            current, depth = queue.popleft()
            if max_depth is not None and depth >= max_depth:
                continue
            source_schema, source_table, source_column = current
            rows = conn.execute(
                """
                SELECT DISTINCT target_schema, target_table, target_column
                FROM lineage_edge
                WHERE source_schema = ? AND source_table = ? AND source_column = ?
                """,
                [source_schema, source_table, source_column],
            ).fetchall()
            for row in rows:
                target = (row[0], row[1], row[2])
                relation_rows.append(
                    [depth + 1, compact_identifier(current), compact_identifier(target)]
                )
                downstream_nodes.add(target)
                if target not in visited:
                    visited.add(target)
                    queue.append((target, depth + 1))

    relation_rows.sort(key=lambda item: (item[0], item[1], item[2]))
    ordered_nodes = sorted(downstream_nodes)
    return relation_rows, ordered_nodes
