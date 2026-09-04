from __future__ import annotations

from datetime import datetime
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from pywebio.output import put_file, put_markdown


def normalize_export_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sanitize_export_fragment(value: str) -> str:
    text = normalize_export_value(value).replace(".", "_").replace(" ", "_")
    return "".join(ch for ch in text if ch.isalnum() or ch in ("_", "-")) or "result"


def create_export_filename(prefix: str, name_parts: list[str], suffix: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_parts = [
        sanitize_export_fragment(part)
        for part in name_parts
        if normalize_export_value(part)
    ]
    body = "_".join([prefix] + safe_parts + [timestamp])
    return f"{body}.{suffix}"


def _ensure_unique_sheet_title(
    raw_title: str, used_titles: set[str], fallback_index: int
) -> str:
    base_title = (normalize_export_value(raw_title) or f"Sheet{fallback_index}")[
        :31
    ] or f"Sheet{fallback_index}"
    candidate = base_title
    suffix = 1
    while candidate in used_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{base_title[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _autosize_sheet(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(normalize_export_value(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _build_sheet_jump_target(sheet_title: str) -> str:
    escaped_title = normalize_export_value(sheet_title)
    if " " in escaped_title:
        return f"#'{escaped_title}'!A1"
    return f"#{escaped_title}!A1"


def build_xlsx_bytes(
    sheet_sections: list[dict],
    add_index_sheet: bool = False,
    index_sheet_title: str = "目录",
) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_titles: set[str] = set()
    rendered_sections: list[tuple[str, dict]] = []
    for index, section in enumerate(sheet_sections, start=1):
        sheet_title = _ensure_unique_sheet_title(
            section.get("title", ""), used_titles, index
        )
        rendered_sections.append((sheet_title, section))

    if add_index_sheet and rendered_sections:
        index_title = _ensure_unique_sheet_title(index_sheet_title, used_titles, 0)
        index_sheet = workbook.create_sheet(title=index_title)
        index_sheet.append(["序号", "Sheet 名称", "最晚出数时间", "跳转"])
        for index, (sheet_title, section) in enumerate(rendered_sections, start=1):
            latest_endtime = normalize_export_value(section.get("index_latest_endtime"))
            index_sheet.append([str(index), sheet_title, latest_endtime, "跳转"])
            link_cell = index_sheet.cell(row=index + 1, column=4)
            link_cell.value = "跳转"
            link_cell.hyperlink = _build_sheet_jump_target(sheet_title)
            link_cell.style = "Hyperlink"
            link_cell.font = Font(color="0563C1", underline="single")
        _autosize_sheet(index_sheet)

    for sheet_title, section in rendered_sections:
        sheet = workbook.create_sheet(title=sheet_title)
        for row in section.get("rows", []):
            sheet.append([normalize_export_value(cell) for cell in row])
        _autosize_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_html_bytes(page_title: str, sheet_sections: list[dict]) -> bytes:
    parts = [
        "<!DOCTYPE html>",
        '<html lang="zh-CN">',
        "<head>",
        '<meta charset="utf-8">',
        f"<title>{escape(page_title)}</title>",
        "<style>",
        "body{font-family:Arial,sans-serif;margin:24px;line-height:1.5;}",
        "h1{font-size:24px;margin-bottom:16px;}",
        "h2{font-size:18px;margin:24px 0 12px;}",
        "table{width:100%;border-collapse:collapse;table-layout:fixed;margin-bottom:20px;}",
        "th,td{border:1px solid #ccc;padding:8px;text-align:left;vertical-align:top;word-break:break-all;}",
        "th{background:#f5f5f5;}",
        "</style>",
        "</head>",
        "<body>",
        f"<h1>{escape(page_title)}</h1>",
    ]

    for section in sheet_sections:
        parts.append(f"<h2>{escape(normalize_export_value(section.get('title')))}</h2>")
        rows = section.get("rows", [])
        if not rows:
            parts.append("<p>无数据</p>")
            continue

        header = rows[0]
        body = rows[1:]
        parts.append("<table>")
        parts.append(
            "<thead><tr>"
            + "".join(
                f"<th>{escape(normalize_export_value(cell))}</th>" for cell in header
            )
            + "</tr></thead>"
        )
        parts.append("<tbody>")
        for row in body:
            rendered_cells = []
            for cell in row:
                cell_text = normalize_export_value(cell)
                if "<a " in cell_text:
                    rendered_cells.append(f"<td>{cell_text}</td>")
                else:
                    rendered_cells.append(f"<td>{escape(cell_text)}</td>")
            parts.append("<tr>" + "".join(rendered_cells) + "</tr>")
        parts.append("</tbody></table>")

    parts.extend(["</body>", "</html>"])
    return "".join(parts).encode("utf-8")


def put_table_exports(
    prefix: str,
    name_parts: list[str],
    page_title: str,
    sheet_sections: list[dict],
    add_index_sheet: bool = False,
    index_sheet_title: str = "目录",
):
    if not sheet_sections:
        return

    put_markdown("### 导出结果")
    put_file(
        create_export_filename(prefix, name_parts, "xlsx"),
        build_xlsx_bytes(
            sheet_sections,
            add_index_sheet=add_index_sheet,
            index_sheet_title=index_sheet_title,
        ),
        "点击下载 Excel（推荐）",
    )
    put_file(
        create_export_filename(prefix, name_parts, "html"),
        build_html_bytes(page_title, sheet_sections),
        "点击下载 HTML",
    )
