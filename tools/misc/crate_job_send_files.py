"""根据作业元数据生成接口说明 Excel。"""

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from pywebio.input import input_group, textarea
from pywebio.output import put_markdown, put_text

from shared.config.metadata import table as metadata_table
from shared.db.gaussdb import select_sql_with_profile
from shared.ui.pywebio_helper import put_table_plus, start_pywebio_app

PROFILE = os.getenv("PYTOOLS_DB_PROFILE", "demo")
WORKSPACE_ROOT = Path(
    os.getenv("PYTOOLS_WORKSPACE_ROOT", "examples/workspace")
).expanduser()
OUTPUT_ROOT = Path(
    os.getenv("PYTOOLS_INTERFACE_OUTPUT_ROOT", "runtime/output")
).expanduser()
DOWNLOAD_ROOT = os.getenv(
    "PYTOOLS_INTERFACE_DOWNLOAD_ROOT", "http://localhost:8500/exports"
).rstrip("/")


def parse_mapping_rules_from_file(file_path: str | Path):
    try:
        data = json.loads(Path(file_path).read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError):
        return None
    return [
        [
            item.get("target_column", ""),
            item.get("mapping_rule", "")
            if len(str(item.get("mapping_rule", ""))) < 30
            else "",
        ]
        for item in data
        if isinstance(item, dict)
    ]


def create_excel(datas: list[list]):
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    border = Border(*(Side(style="thin"),) * 4)
    for data in datas:
        sheet = workbook.create_sheet(
            title=str(data[6]).replace("[DATE]", "")[:31] or "demo"
        )
        for index, title in enumerate(
            ["中文注释", "文件名", "业务逻辑说明", "文件备注", "分隔符", "推送频率"],
            start=1,
        ):
            sheet.cell(index, 1, title).border = border
            sheet.cell(index, 3, data[index - 1]).border = border
        for index, title in enumerate(
            [
                "字段序号",
                "字段名称",
                "中文名称",
                "字段含义",
                "数据产生源系统",
                "字段备注",
            ],
            start=1,
        ):
            sheet.cell(7, index, title).border = border
        for row_index, field in enumerate(data[7] or [], start=8):
            sheet.cell(row_index, 1, row_index - 7)
            sheet.cell(row_index, 2, field[0])
            sheet.cell(row_index, 6, field[1])
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = OUTPUT_ROOT / f"interface_{timestamp}.xlsx"
    workbook.save(output_path)
    put_text(f"生成文件: {output_path.name}")
    put_markdown(
        f'<a href="{DOWNLOAD_ROOT}/{output_path.name}" target="_blank">下载接口说明</a>'
    )


def get_result(job_names: list[str]) -> list[list]:
    jobs_table = metadata_table("jobs", "jobs")
    rows = []
    query = """select job_name, description, event_text, output_path, target_table,
                       calendar, dependency_text
                from __JOBS_TABLE__ where job_name = ?""".replace(
        "__JOBS_TABLE__", jobs_table
    )
    for job_name in job_names:
        result = select_sql_with_profile(PROFILE, query, (job_name,)) or []
        for (
            job,
            description,
            event_text,
            output_path,
            target_table,
            calendar,
            dependency_text,
        ) in result:
            rows.append(
                [
                    "DEMO_INTERFACE",
                    job,
                    description or "",
                    event_text or "",
                    output_path or "",
                    dependency_text or "",
                    target_table or "",
                    "mapping.json",
                    "|",
                    calendar or "",
                    "",
                    target_table or "",
                ]
            )
    return rows


def app():
    info = input_group(
        "生成接口说明文档", [textarea("输入作业名，每行一个", name="job")]
    )
    names = [
        line.strip() for line in str(info.get("job") or "").splitlines() if line.strip()
    ]
    result_rows = get_result(names)
    put_table_plus(
        [["系统", "作业名", "描述", "参数", "输出路径"]]
        + [row[:5] for row in result_rows]
    )
    export_rows = []
    for row in result_rows:
        mapping_file = WORKSPACE_ROOT / row[7]
        export_rows.append(
            [
                row[2],
                Path(row[4]).name or row[4],
                row[5],
                row[6],
                row[8],
                row[9],
                row[4],
                parse_mapping_rules_from_file(mapping_file),
            ]
        )
    if export_rows:
        create_excel(export_rows)


if __name__ == "__main__":
    start_pywebio_app("接口说明文档生成", app)
