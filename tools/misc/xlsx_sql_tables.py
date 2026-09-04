import re
from collections.abc import Mapping
from dataclasses import dataclass
from importlib import import_module
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from shared.lineage.asset_tables import (
    AssetMappingLoadError,
    classify_asset_tables,
    load_asset_plan_map,
)

SQL_HEADER = "SQL脚本"
TABLES_HEADER = "涉及表名"
LEGACY_TABLES_HEADER = "解析表名"
STATUS_HEADER = "解析状态"
ASSET_HEADER = "资产分类"
ASSET_CHECK_FAILED = "校验失败"

STATUS_SUCCESS = "成功"
STATUS_UNRECOGNIZED = "未识别到表"
STATUS_EMPTY = "SQL为空"
STATUS_INVALID = "格式异常"

IGNORED_TABLES = {
    "DATETIME",
    "UTILS.DATEUTILS",
    "DEMO_META.STANDARD_FUNCTIONS",
    "DEMO_META.DATE_FUNCTIONS",
    "DEMO_META.TABLE_FUNCTIONS",
}
IGNORED_IDENTIFIERS = {"LATERAL", "SELECT", "TABLE", "UNNEST", "VALUES"}

_IDENTIFIER_PART = r'(?:"(?:[^"]|"")+"|`[^`]+`|\[[^\]]+\]|[A-Za-z_#$][\w$#]*)'
_QUALIFIED_IDENTIFIER = rf"{_IDENTIFIER_PART}(?:\s*\.\s*{_IDENTIFIER_PART})*"
_TABLE_PATTERNS = [
    re.compile(
        rf"\b(?:FROM|JOIN|USING)\s+(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE
    ),
    re.compile(
        rf"\bINSERT\s+(?:OVERWRITE\s+)?(?:INTO\s+)?(?:TABLE\s+)?(?P<table>{_QUALIFIED_IDENTIFIER})",
        re.IGNORECASE,
    ),
    re.compile(rf"\bUPDATE\s+(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE),
    re.compile(rf"\bMERGE\s+INTO\s+(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE),
    re.compile(rf"\bDELETE\s+FROM\s+(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE),
    re.compile(
        rf"\bCREATE\s+(?:OR\s+REPLACE\s+)?(?:TEMP(?:ORARY)?\s+)?TABLE\s+"
        rf"(?:IF\s+NOT\s+EXISTS\s+)?(?P<table>{_QUALIFIED_IDENTIFIER})",
        re.IGNORECASE,
    ),
    re.compile(rf"\bALTER\s+TABLE\s+(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE),
    re.compile(
        rf"\bTRUNCATE\s+(?:TABLE\s+)?(?P<table>{_QUALIFIED_IDENTIFIER})", re.IGNORECASE
    ),
    re.compile(
        rf"\bDROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(?P<table>{_QUALIFIED_IDENTIFIER})",
        re.IGNORECASE,
    ),
]
_CTE_PATTERN = re.compile(
    rf"(?:\bWITH\b|,)\s*(?:RECURSIVE\s+)?(?P<name>{_IDENTIFIER_PART})"
    rf"(?:\s*\([^)]*\))?\s+AS\s*\(",
    re.IGNORECASE,
)
_IDENTIFIER_PART_PATTERN = re.compile(
    r'(?:(?:"(?:[^"]|"")+"|`[^`]+`|\[[^\]]+\]|[A-Za-z_#$][\w$#]*))'
)

_ASSET_HIT_FILL = PatternFill(fill_type="solid", fgColor="FCE8E6")
_ASSET_HIT_FONT = Font(color="C00000")
_ASSET_FAILURE_FILL = PatternFill(fill_type="solid", fgColor="FFF4CC")
_ASSET_FAILURE_FONT = Font(color="9C6500")
_ASSET_CLEAR_FILL = PatternFill(fill_type=None)
_ASSET_CLEAR_FONT = Font(color=None)


@dataclass(frozen=True)
class ParseResult:
    content: bytes
    sheet_title: str
    successful_rows: int
    unrecognized_rows: int
    empty_rows: int
    invalid_rows: int
    unique_tables: tuple[str, ...]
    asset_related_rows: int
    asset_check_failed: bool

    @property
    def parsed_rows(self) -> int:
        """兼容旧调用方对成功解析行数的读取。"""
        return self.successful_rows


def _mask_sql_comments_and_literals(sql: str) -> str:
    """用空格遮盖注释和单引号字符串，同时保留标识符及文本位置。"""
    chars = list(sql)
    index = 0
    length = len(chars)
    while index < length:
        if sql.startswith("--", index):
            end = sql.find("\n", index + 2)
            end = length if end == -1 else end
            chars[index:end] = " " * (end - index)
            index = end
            continue
        if sql.startswith("/*", index):
            end_marker = sql.find("*/", index + 2)
            end = length if end_marker == -1 else end_marker + 2
            for position in range(index, end):
                if chars[position] not in "\r\n":
                    chars[position] = " "
            index = end
            continue
        if chars[index] == "'":
            start = index
            index += 1
            while index < length:
                if chars[index] == "'":
                    if index + 1 < length and chars[index + 1] == "'":
                        index += 2
                        continue
                    index += 1
                    break
                if chars[index] == "\\" and index + 1 < length:
                    index += 2
                    continue
                index += 1
            for position in range(start, index):
                if chars[position] not in "\r\n":
                    chars[position] = " "
            continue
        if chars[index] in {'"', "`", "["}:
            closing = "]" if chars[index] == "[" else chars[index]
            index += 1
            while index < length:
                if chars[index] == closing:
                    if (
                        closing == '"'
                        and index + 1 < length
                        and chars[index + 1] == '"'
                    ):
                        index += 2
                        continue
                    index += 1
                    break
                index += 1
            continue
        index += 1
    return "".join(chars)


def _normalize_identifier(raw_identifier: str) -> str:
    parts = []
    for match in _IDENTIFIER_PART_PATTERN.finditer(raw_identifier):
        part = match.group(0).strip()
        if part.startswith("[") and part.endswith("]"):
            part = part[1:-1]
        elif part.startswith("`") and part.endswith("`"):
            part = part[1:-1].replace("``", "`")
        elif part.startswith('"') and part.endswith('"'):
            part = part[1:-1].replace('""', '"')
        parts.append(part)
    return ".".join(parts).upper()


def extract_sql_tables(sql: object) -> list[str]:
    """提取 SQL 的来源表、目标表和常见 DDL 操作表。"""
    if not isinstance(sql, str) or not sql.strip():
        return []

    sanitized_sql = _mask_sql_comments_and_literals(sql)
    cte_names = {
        _normalize_identifier(match.group("name"))
        for match in _CTE_PATTERN.finditer(sanitized_sql)
    }
    tables: set[str] = set()
    for pattern in _TABLE_PATTERNS:
        for match in pattern.finditer(sanitized_sql):
            table_name = _normalize_identifier(match.group("table"))
            if (
                not table_name
                or table_name in IGNORED_TABLES
                or table_name in IGNORED_IDENTIFIERS
            ):
                continue
            if "." not in table_name and table_name in cte_names:
                continue
            tables.add(table_name)
    return sorted(tables)


def _header_text(value: object) -> str:
    return value.strip() if isinstance(value, str) else ""


def _find_unique_header(sheet, header: str) -> int:
    matches = [
        column_index
        for column_index in range(1, sheet.max_column + 1)
        if _header_text(sheet.cell(row=1, column=column_index).value) == header
    ]
    if not matches:
        available = [
            _header_text(sheet.cell(row=1, column=column_index).value)
            for column_index in range(1, sheet.max_column + 1)
            if _header_text(sheet.cell(row=1, column=column_index).value)
        ]
        found = "、".join(available) if available else "无有效表头"
        raise ValueError(f"第一个 Sheet 的首行未找到“{header}”表头；当前表头：{found}")
    if len(matches) > 1:
        columns = "、".join(
            sheet.cell(row=1, column=index).column_letter for index in matches
        )
        raise ValueError(f"第一个 Sheet 的首行存在重复“{header}”表头，位置：{columns}")
    return matches[0]


def _find_header_columns(sheet, header: str) -> list[int]:
    return [
        column_index
        for column_index in range(1, sheet.max_column + 1)
        if _header_text(sheet.cell(row=1, column=column_index).value) == header
    ]


def _resolve_output_column(sheet, header: str, legacy_header: str | None = None) -> int:
    matches = _find_header_columns(sheet, header)
    if len(matches) > 1:
        raise ValueError(f"第一个 Sheet 的首行存在重复“{header}”结果列")
    if matches:
        return matches[0]

    if legacy_header:
        legacy_matches = _find_header_columns(sheet, legacy_header)
        if len(legacy_matches) > 1:
            raise ValueError(f"第一个 Sheet 的首行存在重复“{legacy_header}”结果列")
        if legacy_matches:
            sheet.cell(row=1, column=legacy_matches[0], value=header)
            return legacy_matches[0]

    return sheet.max_column + 1


def _style_output_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _row_has_input_data(sheet, row_index: int, input_columns: list[int]) -> bool:
    for column_index in input_columns:
        value = sheet.cell(row=row_index, column=column_index).value
        if value is not None and (not isinstance(value, str) or value.strip()):
            return True
    return False


def _write_asset_result(cell, categories: tuple[str, ...], check_failed: bool):
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if check_failed:
        cell.value = ASSET_CHECK_FAILED
        cell.fill = _ASSET_FAILURE_FILL
        cell.font = _ASSET_FAILURE_FONT
    elif categories:
        cell.value = "\n".join(categories)
        cell.fill = _ASSET_HIT_FILL
        cell.font = _ASSET_HIT_FONT
    else:
        cell.value = ""
        cell.fill = _ASSET_CLEAR_FILL
        cell.font = _ASSET_CLEAR_FONT


def parse_workbook(
    workbook_content: bytes,
    asset_plan_map: Mapping[str, str] | None = None,
    asset_check_failed: bool = False,
) -> ParseResult:
    """解析第一个 Sheet 首行“SQL脚本”列，并追加表名及解析状态。"""
    if not workbook_content:
        raise ValueError("上传的 Excel 文件为空")

    try:
        workbook = load_workbook(BytesIO(workbook_content))
    except Exception as exc:
        raise ValueError(
            f"无法读取 Excel 文件，请确认文件是有效的 .xlsx：{exc}"
        ) from exc

    sheet = workbook.worksheets[0]
    sql_column = _find_unique_header(sheet, SQL_HEADER)
    original_max_column = sheet.max_column
    tables_column = _resolve_output_column(sheet, TABLES_HEADER, LEGACY_TABLES_HEADER)
    tables_header_cell = sheet.cell(row=1, column=tables_column, value=TABLES_HEADER)
    status_column = _resolve_output_column(sheet, STATUS_HEADER)
    status_header_cell = sheet.cell(row=1, column=status_column, value=STATUS_HEADER)
    asset_column = _resolve_output_column(sheet, ASSET_HEADER)
    asset_header_cell = sheet.cell(row=1, column=asset_column, value=ASSET_HEADER)
    _style_output_header(tables_header_cell)
    _style_output_header(status_header_cell)
    _style_output_header(asset_header_cell)

    excluded_columns = {tables_column, status_column, asset_column}
    input_columns = [
        column_index
        for column_index in range(1, original_max_column + 1)
        if column_index not in excluded_columns
    ]
    counts = {
        STATUS_SUCCESS: 0,
        STATUS_UNRECOGNIZED: 0,
        STATUS_EMPTY: 0,
        STATUS_INVALID: 0,
    }
    unique_tables: set[str] = set()
    asset_related_rows = 0
    plan_by_table = asset_plan_map or {}
    for row_index in range(2, sheet.max_row + 1):
        tables_cell: Any = sheet.cell(row=row_index, column=tables_column)
        status_cell: Any = sheet.cell(row=row_index, column=status_column)
        asset_cell: Any = sheet.cell(row=row_index, column=asset_column)
        if not _row_has_input_data(sheet, row_index, input_columns):
            tables_cell.value = ""
            status_cell.value = ""
            _write_asset_result(asset_cell, (), False)
            continue

        sql_cell: Any = sheet.cell(row=row_index, column=sql_column)
        sql = sql_cell.value
        if sql is None or (isinstance(sql, str) and not sql.strip()):
            tables = []
            status = STATUS_EMPTY
        elif not isinstance(sql, str) or sql_cell.data_type == "f":
            tables = []
            status = STATUS_INVALID
        else:
            tables = extract_sql_tables(sql)
            status = STATUS_SUCCESS if tables else STATUS_UNRECOGNIZED

        tables_cell.value = "\n".join(tables)
        tables_cell.alignment = Alignment(wrap_text=True, vertical="top")
        status_cell.value = status
        status_cell.alignment = Alignment(vertical="top")
        asset_categories = classify_asset_tables(tables, plan_by_table)
        _write_asset_result(asset_cell, asset_categories, asset_check_failed)
        if asset_categories and not asset_check_failed:
            asset_related_rows += 1
        counts[status] += 1
        unique_tables.update(tables)

    sheet.column_dimensions[tables_header_cell.column_letter].width = 45
    sheet.column_dimensions[status_header_cell.column_letter].width = 14
    sheet.column_dimensions[asset_header_cell.column_letter].width = 20
    sheet_title = sheet.title
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return ParseResult(
        content=output.getvalue(),
        sheet_title=sheet_title,
        successful_rows=counts[STATUS_SUCCESS],
        unrecognized_rows=counts[STATUS_UNRECOGNIZED],
        empty_rows=counts[STATUS_EMPTY],
        invalid_rows=counts[STATUS_INVALID],
        unique_tables=tuple(sorted(unique_tables)),
        asset_related_rows=asset_related_rows,
        asset_check_failed=asset_check_failed,
    )


def parse_workbook_with_asset_lookup(workbook_content: bytes) -> ParseResult:
    try:
        plan_by_table = load_asset_plan_map()
    except AssetMappingLoadError:
        return parse_workbook(workbook_content, asset_check_failed=True)
    return parse_workbook(workbook_content, asset_plan_map=plan_by_table)


def build_output_filename(input_filename: str) -> str:
    stem = Path(input_filename or "SQL脚本.xlsx").stem or "SQL脚本"
    return f"{stem}_SQL表名解析.xlsx"


def main():
    pywebio_input = import_module("pywebio.input")
    pywebio_output = import_module("pywebio.output")
    file_upload = pywebio_input.file_upload
    put_file = pywebio_output.put_file
    put_markdown = pywebio_output.put_markdown

    from shared.ui.pywebio_helper import safe_put_error

    put_markdown("## Excel SQL 表名解析")
    put_markdown(
        "上传 Excel 后，工具会读取第一个 Sheet 首行的 **SQL脚本** 列，并生成涉及表名和解析状态。"
    )
    uploaded = file_upload("请选择 Excel 文件", accept=".xlsx", required=True)
    try:
        result = parse_workbook_with_asset_lookup(uploaded["content"])
    except Exception as exc:
        safe_put_error(exc)
        return

    put_markdown(
        f"处理完成：成功 **{result.successful_rows}** 行，"
        f"未识别 **{result.unrecognized_rows}** 行，"
        f"SQL 为空 **{result.empty_rows}** 行，"
        f"格式异常 **{result.invalid_rows}** 行，"
        f"共识别 **{len(result.unique_tables)}** 张表，"
        f"资产分类命中 **{result.asset_related_rows}** 行。"
    )
    if result.asset_check_failed:
        put_markdown(
            "> ⚠️ demo 资产映射加载失败，Excel 中已标记为“校验失败”；SQL 表名解析结果不受影响。"
        )
    put_file(
        build_output_filename(uploaded.get("filename", "SQL脚本.xlsx")),
        result.content,
        "下载解析后的 Excel",
    )


if __name__ == "__main__":
    from shared.ui.pywebio_helper import start_pywebio_app

    start_pywebio_app("Excel SQL 表名解析", main)
